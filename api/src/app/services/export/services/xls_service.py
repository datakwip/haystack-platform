from openpyxl import Workbook
from app.services.export.services import entity_service
import io
def removeLoadedData(data, loaded_data):
    res = []
    for d in data:
        if d["db_id"] not in loaded_data:
            loaded_data[d["db_id"]] = d["db_id"]
            res.append(d)
    return res


def getSheetName(tags, tag_hierarchy):
    sheetName = ""
    tagList = tags.split(" ")
    for tag in tagList:
        if (tag != tag_hierarchy[tag]):
            sheetName += "{}({}) ".format(tag, tag_hierarchy[tag])
        else:
            sheetName += "{} ".format(tag)
    return sheetName.strip()

def fillInfoTab(infoSheet, hierarchy):
    i = 1
    for entity_type in hierarchy:
        infoSheet.cell(row=1, column=i).value = entity_type[0]
        j = 2
        for tags in entity_type[1]["real_tags"]:
            infoSheet.cell(row=j, column=i).value = tags
            j+=1
        i += 1
def createXLSFile(hierarchy, conn, org_id, tags, entity_refs, schema, tag_hierarchy, save: bool=True):
    wb = Workbook()
    loaded_data = {}
    infoSheet = wb.active
    infoSheet.title = "info"
    fillInfoTab(infoSheet, hierarchy)
    for entity_type in hierarchy:
        #sheetName =getSheetName(entity_type[0], tag_hierarchy)
        ws = wb.create_sheet(title=entity_type[0])
        columns, tempData = entity_service.get_entities_by_type(entity_type,conn, org_id, tags, entity_refs, schema)
        if len(tempData) > 0:
            data = removeLoadedData(tempData, loaded_data)
            for idx, x in enumerate(columns):
                ws.cell(row=1, column=idx+1).value = x
            j=1
            for idx, x in enumerate(data):
               j +=1
               for k in x.keys():
                   i = columns.index(k)
                   ws.cell(row=j, column=i + 1).value = x[k]


    # Save the file
    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
